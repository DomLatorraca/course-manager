from __future__ import annotations

from dataclasses import dataclass, field
from hashlib import sha1
from pathlib import Path
import re
from typing import Any

from sqlalchemy import func, select
from sqlalchemy.orm import Session

from app.models.course import Course, CourseStatus
from app.models.enrollment import Enrollment
from app.models.student import Student
from app.services.excel_course_service import get_sheet_name, list_excel_courses
from app.services.student_service import STUDENT_EMAIL_DOMAIN
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException


@dataclass(frozen=True)
class ExcelCourseImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    students_created: int = 0
    students_updated: int = 0
    enrollments_created: int = 0
    enrollments_updated: int = 0
    enrollments_skipped: int = 0
    errors: list[str] = field(default_factory=list)


def import_excel_courses_to_db(
    db: Session,
    path: Path | None = None,
    sheet_name: str | None = None,
    update_existing: bool = True,
    dry_run: bool = False,
    include_participants: bool = True,
) -> ExcelCourseImportResult:
    created = 0
    updated = 0
    skipped = 0
    students_created = 0
    students_updated = 0
    enrollments_created = 0
    enrollments_updated = 0
    enrollments_skipped = 0
    errors: list[str] = []

    excel_courses = sorted(list_excel_courses(path=path, sheet_name=sheet_name), key=lambda course: course.id)
    excel_course_titles = {course.title.strip().lower() for course in excel_courses}
    for excel_course in excel_courses:
        title = excel_course.title.strip()
        if not title:
            skipped += 1
            errors.append(f"Corso Excel ID {excel_course.id}: titolo mancante.")
            continue

        existing = db.scalar(select(Course).where(func.lower(Course.title) == title.lower()))
        if existing:
            if not update_existing:
                skipped += 1
                continue
            updated += 1
            if not dry_run:
                _copy_excel_course_to_model(existing, excel_course)
                db.add(existing)
            continue

        created += 1
        if not dry_run:
            course = Course()
            _copy_excel_course_to_model(course, excel_course)
            db.add(course)

    if not dry_run:
        db.flush()

    if include_participants:
        for participant in _extract_calendar_participants(path, sheet_name):
            course = _find_course_by_title(db, participant.course_title)
            if not course:
                if dry_run and participant.course_title.lower() in excel_course_titles:
                    course_would_exist = True
                else:
                    course_would_exist = False
                if course_would_exist:
                    pass
                else:
                    errors.append(f"Partecipante '{participant.full_name}': corso '{participant.course_title}' non trovato nel database.")
                    continue

            if dry_run and not course:
                student = _find_imported_student(db, participant.full_name)
                if student:
                    students_updated += 1
                else:
                    students_created += 1
                enrollments_created += 1
                continue

            if not course:
                errors.append(f"Partecipante '{participant.full_name}': corso '{participant.course_title}' non trovato nel database.")
                continue

            email = _generated_student_email(participant.full_name)
            student = _find_imported_student(db, participant.full_name)
            first_name, last_name = _split_full_name(participant.full_name)
            student_notes = ""

            if student:
                if _student_needs_update(student, first_name, last_name, student_notes, email):
                    students_updated += 1
                    if not dry_run:
                        student.first_name = first_name
                        student.last_name = last_name
                        student.email = email
                        student.notes = student_notes
                        db.add(student)
                else:
                    skipped += 1
            else:
                students_created += 1
                if not dry_run:
                    student = Student(
                        first_name=first_name,
                        last_name=last_name,
                        email=email,
                        organization="",
                        notes=student_notes,
                    )
                    db.add(student)
                    db.flush()

            if dry_run:
                enrollments_created += 1
                continue

            enrollment = db.scalar(
                select(Enrollment).where(
                    Enrollment.course_id == course.id,
                    Enrollment.student_id == student.id,
                )
            )
            if enrollment:
                if enrollment.notes != participant.enrollment_notes:
                    enrollment.notes = participant.enrollment_notes
                    db.add(enrollment)
                    enrollments_updated += 1
                else:
                    enrollments_skipped += 1
            else:
                db.add(
                    Enrollment(
                        course_id=course.id,
                        student_id=student.id,
                        notes=participant.enrollment_notes,
                    )
                )
                enrollments_created += 1

    if dry_run:
        db.rollback()
    else:
        db.commit()

    return ExcelCourseImportResult(
        created=created,
        updated=updated,
        skipped=skipped,
        students_created=students_created,
        students_updated=students_updated,
        enrollments_created=enrollments_created,
        enrollments_updated=enrollments_updated,
        enrollments_skipped=enrollments_skipped,
        errors=errors,
    )


def _copy_excel_course_to_model(course: Course, excel_course) -> None:
    course.title = excel_course.title
    course.short_description = "" if _is_excel_source_text(excel_course.short_description) else excel_course.short_description
    course.category = "" if _is_excel_sheet_name(excel_course.category) else excel_course.category
    course.duration = excel_course.duration
    course.status = CourseStatus(excel_course.status)


@dataclass(frozen=True)
class CalendarParticipant:
    course_title: str
    full_name: str
    source_sheet: str
    source_row: int
    enrollment_notes: str


def _extract_calendar_participants(path: Path | None, sheet_name: str | None) -> list[CalendarParticipant]:
    workbook_path = Path(path).expanduser().resolve() if path else None
    if workbook_path is None:
        from app.services.excel_course_service import get_excel_path

        workbook_path = get_excel_path()

    try:
        workbook = load_workbook(workbook_path, data_only=True)
    except (InvalidFileException, OSError) as exc:
        raise ValueError(f"Impossibile leggere il file Excel: {exc}") from exc

    course_sheet_name = sheet_name or get_sheet_name()
    participants: list[CalendarParticipant] = []
    for worksheet in workbook.worksheets:
        if worksheet.title == course_sheet_name:
            continue

        title_rows = _find_course_title_rows(worksheet)
        for index, (title_row, course_title) in enumerate(title_rows):
            next_title_row = title_rows[index + 1][0] if index + 1 < len(title_rows) else worksheet.max_row + 1
            for row_number in range(title_row + 4, next_title_row):
                full_name = _clean_text(worksheet.cell(row=row_number, column=1).value)
                if not full_name or _is_non_participant_label(full_name):
                    continue
                notes = _calendar_notes_for_row(worksheet.title, row_number, worksheet[row_number][1:])
                participants.append(
                    CalendarParticipant(
                        course_title=course_title,
                        full_name=full_name,
                        source_sheet=worksheet.title,
                        source_row=row_number,
                        enrollment_notes=notes,
                    )
                )

    return participants


def _find_course_title_rows(worksheet) -> list[tuple[int, str]]:
    rows: list[tuple[int, str]] = []
    for row in worksheet.iter_rows():
        for cell in row:
            value = _clean_text(cell.value)
            if value and "CORSO" in value.upper():
                rows.append((cell.row, " ".join(value.split())))
                break
    return rows


def _calendar_notes_for_row(sheet_title: str, row_number: int, cells: tuple[Any, ...]) -> str:
    values: dict[str, int] = {}
    for cell in cells:
        value = _clean_text(cell.value)
        if not value:
            continue
        if _looks_like_number(value):
            continue
        normalized = " ".join(value.split())
        values[normalized] = values.get(normalized, 0) + 1

    if not values:
        return ""

    details = ", ".join(f"{label} ({count})" for label, count in sorted(values.items()))
    return f"Valori calendario: {details}."


def _find_course_by_title(db: Session, title: str) -> Course | None:
    return db.scalar(select(Course).where(func.lower(Course.title) == title.lower()))


def _student_needs_update(student: Student, first_name: str, last_name: str, notes: str, email: str) -> bool:
    return student.first_name != first_name or student.last_name != last_name or student.notes != notes or student.email != email


def _split_full_name(full_name: str) -> tuple[str, str]:
    parts = full_name.strip().split()
    if not parts:
        return "Sconosciuto", "Sconosciuto"
    if len(parts) == 1:
        return parts[0], parts[0]

    surname_words = 1
    first = parts[0].lower().replace("'", "")
    if first in {"da", "de", "del", "della", "di", "dal", "dello"} and len(parts) > 2:
        surname_words = 2

    last_name = " ".join(parts[:surname_words])
    first_name = " ".join(parts[surname_words:])
    return first_name, last_name


def _generated_student_email(full_name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", full_name.lower()).strip("-") or "iscritto"
    digest = sha1(full_name.strip().casefold().encode("utf-8")).hexdigest()[:10]
    return f"excel-{slug}-{digest}@{STUDENT_EMAIL_DOMAIN}"


def _legacy_generated_student_email(full_name: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", full_name.lower()).strip("-") or "iscritto"
    digest = sha1(full_name.strip().casefold().encode("utf-8")).hexdigest()[:10]
    return f"excel-{slug}-{digest}@course-manager.invalid"


def _find_imported_student(db: Session, full_name: str) -> Student | None:
    current_email = _generated_student_email(full_name)
    legacy_email = _legacy_generated_student_email(full_name)
    current = db.scalar(select(Student).where(Student.email == current_email))
    if current:
        return current
    return db.scalar(select(Student).where(Student.email == legacy_email))


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _looks_like_number(value: str) -> bool:
    try:
        float(value)
        return True
    except ValueError:
        return False


def _is_non_participant_label(value: str) -> bool:
    normalized = value.strip().upper()
    return normalized in {"LEGENDA", "GRIMANI", "DOCENTE ESTERNO", "ONLINE", "PRESENZA"}


def _is_excel_source_text(value: str) -> bool:
    return _clean_text(value).lower().startswith("importato dal foglio")


def _is_excel_sheet_name(value: str) -> bool:
    return bool(re.fullmatch(r"foglio\s*\d*", _clean_text(value), flags=re.IGNORECASE))
