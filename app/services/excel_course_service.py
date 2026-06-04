from __future__ import annotations

from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet

from app.config import get_settings


HEADERS = [
    "id",
    "titolo",
    "descrizione_breve",
    "categoria",
    "durata",
    "stato",
    "data_creazione",
    "data_aggiornamento",
]
VALID_STATUSES = {"attivo", "archiviato"}


class ExcelCourseError(ValueError):
    pass


@dataclass(frozen=True)
class ExcelCourse:
    id: int
    title: str
    short_description: str
    category: str
    duration: str
    status: str
    created_at: datetime
    updated_at: datetime


@dataclass(frozen=True)
class ExcelCoursePayload:
    title: str
    short_description: str = ""
    category: str = ""
    duration: str = ""
    status: str = "attivo"


def get_excel_path() -> Path:
    return get_settings().excel_courses_path


def get_sheet_name() -> str:
    return get_settings().excel_courses_sheet_name


def list_excel_courses(q: str = "", category: str = "", status: str = "", path: Path | None = None, sheet_name: str | None = None) -> list[ExcelCourse]:
    workbook, worksheet, should_save = _load_course_sheet(path, sheet_name)
    if should_save:
        _save_workbook(workbook, _resolve_path(path))

    courses = [_course_from_row(row) for row in worksheet.iter_rows(min_row=2, values_only=True) if _row_has_course(row)]
    courses = [course for course in courses if course is not None]

    if q:
        needle = q.casefold()
        courses = [
            course
            for course in courses
            if needle in course.title.casefold() or needle in course.short_description.casefold()
        ]
    if category:
        needle = category.casefold()
        courses = [course for course in courses if needle in course.category.casefold()]
    if status:
        courses = [course for course in courses if course.status == status]

    return sorted(courses, key=lambda course: (course.updated_at, course.id), reverse=True)


def get_excel_course(course_id: int, path: Path | None = None, sheet_name: str | None = None) -> ExcelCourse | None:
    workbook, worksheet, should_save = _load_course_sheet(path, sheet_name)
    if should_save:
        _save_workbook(workbook, _resolve_path(path))

    row_number = _find_row_number(worksheet, course_id)
    if row_number is None:
        return None
    return _course_from_row([cell.value for cell in worksheet[row_number]])


def create_excel_course(payload: ExcelCoursePayload, path: Path | None = None, sheet_name: str | None = None) -> ExcelCourse:
    payload = _clean_payload(payload)
    workbook, worksheet, _ = _load_course_sheet(path, sheet_name)
    now = _utcnow()
    course_id = _next_id(worksheet)
    worksheet.append(
        [
            course_id,
            payload.title,
            payload.short_description,
            payload.category,
            payload.duration,
            payload.status,
            now,
            now,
        ]
    )
    _format_sheet(worksheet)
    _save_workbook(workbook, _resolve_path(path))
    created = get_excel_course(course_id, path, sheet_name)
    if not created:
        raise ExcelCourseError("Corso creato ma non riletto dal file Excel.")
    return created


def update_excel_course(course_id: int, payload: ExcelCoursePayload, path: Path | None = None, sheet_name: str | None = None) -> ExcelCourse:
    payload = _clean_payload(payload)
    workbook, worksheet, _ = _load_course_sheet(path, sheet_name)
    row_number = _find_row_number(worksheet, course_id)
    if row_number is None:
        raise ExcelCourseError("Corso non trovato nel file Excel.")

    row = worksheet[row_number]
    row[1].value = payload.title
    row[2].value = payload.short_description
    row[3].value = payload.category
    row[4].value = payload.duration
    row[5].value = payload.status
    if not row[6].value:
        row[6].value = _utcnow()
    row[7].value = _utcnow()
    _format_sheet(worksheet)
    _save_workbook(workbook, _resolve_path(path))

    updated = get_excel_course(course_id, path, sheet_name)
    if not updated:
        raise ExcelCourseError("Corso aggiornato ma non riletto dal file Excel.")
    return updated


def archive_excel_course(course_id: int, path: Path | None = None, sheet_name: str | None = None) -> ExcelCourse:
    course = get_excel_course(course_id, path, sheet_name)
    if not course:
        raise ExcelCourseError("Corso non trovato nel file Excel.")
    return update_excel_course(
        course_id,
        ExcelCoursePayload(
            title=course.title,
            short_description=course.short_description,
            category=course.category,
            duration=course.duration,
            status="archiviato",
        ),
        path,
        sheet_name,
    )


def delete_excel_course(course_id: int, path: Path | None = None, sheet_name: str | None = None) -> None:
    workbook, worksheet, _ = _load_course_sheet(path, sheet_name)
    row_number = _find_row_number(worksheet, course_id)
    if row_number is None:
        raise ExcelCourseError("Corso non trovato nel file Excel.")
    worksheet.delete_rows(row_number, 1)
    _format_sheet(worksheet)
    _save_workbook(workbook, _resolve_path(path))


def _resolve_path(path: Path | None) -> Path:
    return Path(path or get_excel_path()).expanduser().resolve()


def _resolve_sheet_name(sheet_name: str | None) -> str:
    return sheet_name or get_sheet_name()


def _load_course_sheet(path: Path | None, sheet_name: str | None) -> tuple[Workbook, Worksheet, bool]:
    workbook_path = _resolve_path(path)
    target_sheet = _resolve_sheet_name(sheet_name)
    should_save = False

    if workbook_path.exists():
        try:
            workbook = load_workbook(workbook_path)
        except (InvalidFileException, OSError) as exc:
            raise ExcelCourseError(f"Impossibile leggere il file Excel: {exc}") from exc
    else:
        workbook = Workbook()
        default = workbook.active
        default.title = target_sheet
        should_save = True

    if target_sheet in workbook.sheetnames:
        worksheet = workbook[target_sheet]
        if _sheet_is_empty(worksheet):
            _initialize_sheet(worksheet)
            should_save = True
    else:
        worksheet = workbook.create_sheet(target_sheet)
        _initialize_sheet(worksheet)
        _seed_courses_from_calendar(workbook, worksheet)
        should_save = True

    if _headers_are_missing(worksheet):
        raise ExcelCourseError(
            f"Il foglio '{target_sheet}' esiste ma non ha le intestazioni attese: {', '.join(HEADERS)}."
        )

    _format_sheet(worksheet)
    return workbook, worksheet, should_save


def _initialize_sheet(worksheet: Worksheet) -> None:
    worksheet.delete_rows(1, worksheet.max_row)
    worksheet.append(HEADERS)


def _sheet_is_empty(worksheet: Worksheet) -> bool:
    return worksheet.max_row == 1 and all(cell.value in (None, "") for cell in worksheet[1])


def _headers_are_missing(worksheet: Worksheet) -> bool:
    values = [cell.value for cell in worksheet[1][: len(HEADERS)]]
    if all(value in (None, "") for value in values):
        return True
    return [str(value or "").strip() for value in values] != HEADERS


def _seed_courses_from_calendar(workbook: Workbook, target: Worksheet) -> bool:
    now = _utcnow()
    seen: set[str] = set()
    next_id = 1

    for worksheet in workbook.worksheets:
        if worksheet is target:
            continue
        for row in worksheet.iter_rows():
            for cell in row:
                value = _clean_text(cell.value)
                if not value or "CORSO" not in value.upper():
                    continue
                normalized = " ".join(value.split())
                if normalized.casefold() in seen:
                    continue
                seen.add(normalized.casefold())
                target.append(
                    [
                        next_id,
                        normalized,
                        f"Importato dal foglio {worksheet.title}, riga {cell.row}",
                        worksheet.title,
                        "",
                        "attivo",
                        now,
                        now,
                    ]
                )
                next_id += 1
                break

    return next_id > 1


def _save_workbook(workbook: Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile(delete=False, suffix=".xlsx", dir=path.parent) as tmp:
        tmp_path = Path(tmp.name)
    try:
        workbook.save(tmp_path)
        tmp_path.replace(path)
    except OSError as exc:
        raise ExcelCourseError(f"Impossibile salvare il file Excel. Verificare che non sia aperto in Excel: {exc}") from exc
    finally:
        if tmp_path.exists():
            tmp_path.unlink(missing_ok=True)


def _format_sheet(worksheet: Worksheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F2933")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    widths = {
        "A": 10,
        "B": 34,
        "C": 42,
        "D": 20,
        "E": 18,
        "F": 14,
        "G": 22,
        "H": 22,
    }
    for column, width in widths.items():
        worksheet.column_dimensions[column].width = width
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:H{max(worksheet.max_row, 1)}"


def _row_has_course(row: tuple[Any, ...]) -> bool:
    if not row:
        return False
    return row[0] not in (None, "") and row[1] not in (None, "")


def _course_from_row(row: tuple[Any, ...] | list[Any]) -> ExcelCourse | None:
    try:
        course_id = int(row[0])
    except (TypeError, ValueError):
        return None

    return ExcelCourse(
        id=course_id,
        title=_clean_text(row[1]),
        short_description=_clean_text(row[2]),
        category=_clean_text(row[3]),
        duration=_clean_text(row[4]),
        status=_normalize_status(row[5]),
        created_at=_parse_datetime(row[6]),
        updated_at=_parse_datetime(row[7]),
    )


def _find_row_number(worksheet: Worksheet, course_id: int) -> int | None:
    for row_number in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row=row_number, column=1).value
        try:
            if int(value) == course_id:
                return row_number
        except (TypeError, ValueError):
            continue
    return None


def _next_id(worksheet: Worksheet) -> int:
    current_ids: list[int] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        try:
            current_ids.append(int(row[0]))
        except (TypeError, ValueError):
            continue
    return max(current_ids, default=0) + 1


def _clean_payload(payload: ExcelCoursePayload) -> ExcelCoursePayload:
    title = _clean_text(payload.title)
    status = _normalize_status(payload.status)
    if not title:
        raise ExcelCourseError("Il titolo del corso è obbligatorio.")
    return ExcelCoursePayload(
        title=title,
        short_description=_clean_text(payload.short_description),
        category=_clean_text(payload.category),
        duration=_clean_text(payload.duration),
        status=status,
    )


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_status(value: Any) -> str:
    status = _clean_text(value).lower() or "attivo"
    if status not in VALID_STATUSES:
        raise ExcelCourseError("Stato corso non valido.")
    return status


def _parse_datetime(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    if isinstance(value, str) and value:
        try:
            return datetime.fromisoformat(value)
        except ValueError:
            pass
    return _utcnow()


def _utcnow() -> datetime:
    return datetime.now(UTC).replace(tzinfo=None)
