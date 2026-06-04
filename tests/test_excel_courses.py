from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from sqlalchemy import select

from app.models.course import Course
from app.models.enrollment import Enrollment
from app.models.student import Student
from app.services.excel_course_service import (
    ExcelCoursePayload,
    archive_excel_course,
    create_excel_course,
    delete_excel_course,
    list_excel_courses,
    update_excel_course,
)
from app.services.excel_import_service import import_excel_courses_to_db


def test_excel_courses_sheet_is_seeded_from_calendar_titles(tmp_path: Path):
    path = tmp_path / "formazione.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Foglio3"
    worksheet["B5"] = "CORSO FULL STACK 2025"
    worksheet["B14"] = "CORSO AI 2025 - 2026"
    workbook.save(path)

    courses = list_excel_courses(path=path, sheet_name="Corsi")

    assert [course.title for course in courses] == ["CORSO AI 2025 - 2026", "CORSO FULL STACK 2025"]
    saved = load_workbook(path)
    assert "Foglio3" in saved.sheetnames
    assert "Corsi" in saved.sheetnames
    assert saved["Corsi"]["B2"].value == "CORSO FULL STACK 2025"


def test_excel_course_crud_uses_workbook_rows(tmp_path: Path):
    path = tmp_path / "corsi.xlsx"

    created = create_excel_course(
        ExcelCoursePayload(
            title="Python base",
            short_description="Corso introduttivo",
            category="Sviluppo",
            duration="16h",
        ),
        path=path,
        sheet_name="Corsi",
    )

    assert created.id == 1
    assert created.status == "attivo"

    updated = update_excel_course(
        created.id,
        ExcelCoursePayload(
            title="Python avanzato",
            short_description="Corso aggiornato",
            category="Sviluppo",
            duration="24h",
            status="attivo",
        ),
        path=path,
        sheet_name="Corsi",
    )
    assert updated.title == "Python avanzato"
    assert updated.duration == "24h"

    archived = archive_excel_course(created.id, path=path, sheet_name="Corsi")
    assert archived.status == "archiviato"

    assert len(list_excel_courses(status="archiviato", path=path, sheet_name="Corsi")) == 1

    delete_excel_course(created.id, path=path, sheet_name="Corsi")
    assert list_excel_courses(path=path, sheet_name="Corsi") == []


def test_import_excel_courses_populates_db_from_calendar_titles(db, tmp_path: Path):
    path = tmp_path / "formazione.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Foglio3"
    worksheet["B5"] = "CORSO FULL STACK 2025"
    worksheet["B7"] = "MAGGIO 2025"
    worksheet["B8"] = 2
    worksheet["A9"] = "Rossi Mario"
    worksheet["B9"] = "GRIMANI"
    worksheet["C9"] = "DOCENTE ESTERNO"
    workbook.save(path)

    result = import_excel_courses_to_db(db, path=path, sheet_name="Corsi")

    assert result.created == 1
    assert result.updated == 0
    assert result.students_created == 1
    assert result.enrollments_created == 1
    course = db.scalar(select(Course).where(Course.title == "CORSO FULL STACK 2025"))
    assert course is not None
    assert course.category == "Foglio3"
    student = db.scalar(select(Student).where(Student.last_name == "Rossi", Student.first_name == "Mario"))
    assert student is not None
    enrollment = db.scalar(select(Enrollment).where(Enrollment.course_id == course.id, Enrollment.student_id == student.id))
    assert enrollment is not None
    assert "GRIMANI" in enrollment.notes
    assert "DOCENTE ESTERNO" in enrollment.notes


def test_import_excel_courses_updates_existing_db_course(db, tmp_path: Path):
    path = tmp_path / "corsi.xlsx"
    create_excel_course(
        ExcelCoursePayload(
            title="Python base",
            short_description="Prima descrizione",
            category="Sviluppo",
            duration="16h",
        ),
        path=path,
        sheet_name="Corsi",
    )

    first_result = import_excel_courses_to_db(db, path=path, sheet_name="Corsi")
    assert first_result.created == 1

    update_excel_course(
        1,
        ExcelCoursePayload(
            title="Python base",
            short_description="Descrizione aggiornata",
            category="Academy",
            duration="24h",
        ),
        path=path,
        sheet_name="Corsi",
    )
    second_result = import_excel_courses_to_db(db, path=path, sheet_name="Corsi")

    assert second_result.created == 0
    assert second_result.updated == 1
    courses = db.scalars(select(Course).where(Course.title == "Python base")).all()
    assert len(courses) == 1
    assert courses[0].short_description == "Descrizione aggiornata"
    assert courses[0].category == "Academy"
    assert courses[0].duration == "24h"


def test_import_excel_courses_is_idempotent_for_participants(db, tmp_path: Path):
    path = tmp_path / "formazione.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Foglio3"
    worksheet["B5"] = "CORSO AI 2025"
    worksheet["B7"] = "GIUGNO 2025"
    worksheet["B8"] = 10
    worksheet["A9"] = "Di Dio Giovanni Paolo"
    worksheet["B9"] = "ONLINE"
    workbook.save(path)

    first_result = import_excel_courses_to_db(db, path=path, sheet_name="Corsi")
    second_result = import_excel_courses_to_db(db, path=path, sheet_name="Corsi")

    assert first_result.students_created == 1
    assert first_result.enrollments_created == 1
    assert second_result.students_created == 0
    assert second_result.enrollments_created == 0
    assert second_result.enrollments_skipped == 1
    students = db.scalars(select(Student)).all()
    enrollments = db.scalars(select(Enrollment)).all()
    assert len(students) == 1
    assert len(enrollments) == 1
    assert students[0].last_name == "Di Dio"
    assert students[0].first_name == "Giovanni Paolo"
